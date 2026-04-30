"""
Parts Orders Module — Purchase Requisition Form
Mirrors the layout of Purchase Req_MASTER.xlsx and writes submitted
entries back into the template for saving / distribution.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import date
import os
import re

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Template location (same folder as this file) ─────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(_HERE, "Purchase Req_MASTER.xlsx")

SHEET_NAME = "Commercial Parts Purchase Req"

# ── Auto-populated constants ───────────────────────────────────────────────────
SHIP_TO   = "Airbus c/o AIT 320 Airbus Way A220 FAL Building Mobile AL 36615"
ISSUER    = "Bradley Drake"
JOB_ID_OPTIONS = ["1000", "1001", "SOTC"]

# ── Column headers for line items (mirrors row 6 of the spreadsheet) ──────────
ITEM_COLS = [
    "Qty", "Description", "Supplier Name", "Part #", "Quote Link",
    "Unit Price", "Ext. Price\n(auto)", "Date Required",
    "CP Related\n(Y/N)", "If YES,\nCP #", "Quote\nNumber",
    "Approved By", "Leadtime", "New Supplier\n(Y/N)",
    "Reason New Supplier\nIs Needed",
]
# Widths (chars) for each column entry widget
ITEM_WIDTHS = [5, 24, 18, 14, 22, 10, 10, 12, 8, 8, 10, 14, 10, 8, 24]
# Max line-item rows in the template
MAX_ROWS = 23
DEFAULT_ROWS = 5


# ─────────────────────────────────────────────────────────────────────────────
class PartsOrderPanel(tk.Frame):
    """Embeddable Parts Order / Purchase Requisition form."""

    def __init__(self, master):
        super().__init__(master)
        self.configure(bg="#f0f4f8")
        self._rows: list[dict] = []   # list of dicts: {col_name: widget}
        self._build()

    # ── Build UI ──────────────────────────────────────────────────────────────
    def _build(self):
        # ── Header bar
        hdr = tk.Frame(self, bg="#1a3c5e", pady=8)
        hdr.pack(fill="x")
        tk.Label(hdr, text="📦  Parts Orders — Purchase Requisition",
                 bg="#1a3c5e", fg="white",
                 font=("Segoe UI", 14, "bold")).pack(side="left", padx=16)

        # ── Scroll canvas so the whole form is scrollable
        outer = tk.Frame(self, bg="#f0f4f8")
        outer.pack(fill="both", expand=True)
        canvas = tk.Canvas(outer, bg="#f0f4f8", highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        hsb = ttk.Scrollbar(outer, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        canvas.pack(fill="both", expand=True)

        body = tk.Frame(canvas, bg="#f0f4f8", padx=16, pady=12)
        win_id = canvas.create_window((0, 0), window=body, anchor="nw")
        body.bind("<Configure>",
                  lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfig(win_id, width=e.width))
        canvas.bind_all("<MouseWheel>",
            lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))

        self._build_header_section(body)
        self._build_items_section(body)
        self._build_footer_buttons(body)

    # ── Header section ────────────────────────────────────────────────────────
    def _build_header_section(self, parent):
        sec = self._section(parent, "📋  Requisition Header")

        # Row 1: Issuer | Date | Airbus Requestor
        r1 = tk.Frame(sec, bg="#ffffff"); r1.pack(fill="x", pady=4)

        self._lbl(r1, "Issuer:").grid(row=0, column=0, sticky="e", padx=(0, 6))
        self.issuer_var = tk.StringVar(value=ISSUER)
        tk.Entry(r1, textvariable=self.issuer_var, width=22,
                 font=("Segoe UI", 10)).grid(row=0, column=1, padx=(0, 20))

        self._lbl(r1, "Date:").grid(row=0, column=2, sticky="e", padx=(0, 6))
        self.date_var = tk.StringVar(value=date.today().strftime("%m/%d/%Y"))
        tk.Entry(r1, textvariable=self.date_var, width=14,
                 font=("Segoe UI", 10)).grid(row=0, column=3, padx=(0, 20))

        self._lbl(r1, "Airbus Requestor:").grid(row=0, column=4, sticky="e", padx=(0, 6))
        self.requestor_var = tk.StringVar()
        tk.Entry(r1, textvariable=self.requestor_var, width=22,
                 font=("Segoe UI", 10)).grid(row=0, column=5)

        # Row 2: Ship To (read-only) | Job ID dropdown
        r2 = tk.Frame(sec, bg="#ffffff"); r2.pack(fill="x", pady=4)

        self._lbl(r2, "Ship To:").grid(row=0, column=0, sticky="ne", padx=(0, 6))
        ship_lbl = tk.Label(r2, text=SHIP_TO, bg="#eef2f7", fg="#1a3c5e",
                            font=("Segoe UI", 9, "italic"),
                            relief="sunken", padx=6, pady=4,
                            wraplength=500, justify="left", anchor="w")
        ship_lbl.grid(row=0, column=1, sticky="w", padx=(0, 30))

        self._lbl(r2, "Job ID Parent 13382-ABA:").grid(
            row=0, column=2, sticky="e", padx=(0, 6))
        self.job_id_var = tk.StringVar(value=JOB_ID_OPTIONS[0])
        ttk.Combobox(r2, textvariable=self.job_id_var,
                     values=JOB_ID_OPTIONS, width=10,
                     state="readonly", font=("Segoe UI", 10)).grid(
            row=0, column=3, sticky="w")

    # ── Line items section ────────────────────────────────────────────────────
    def _build_items_section(self, parent):
        sec = self._section(parent, "🧾  Line Items")

        # Column header row
        hdr_frame = tk.Frame(sec, bg="#1a3c5e")
        hdr_frame.pack(fill="x", pady=(0, 2))
        tk.Label(hdr_frame, text="#", bg="#1a3c5e", fg="white",
                 font=("Segoe UI", 8, "bold"), width=3).pack(side="left", padx=2)
        for i, (col, w) in enumerate(zip(ITEM_COLS, ITEM_WIDTHS)):
            tk.Label(hdr_frame, text=col, bg="#1a3c5e", fg="white",
                     font=("Segoe UI", 8, "bold"), width=w,
                     wraplength=w * 7, justify="center").pack(side="left", padx=2)

        # Scrollable rows container
        self._items_frame = tk.Frame(sec, bg="#f0f4f8")
        self._items_frame.pack(fill="x")

        for _ in range(DEFAULT_ROWS):
            self._add_row()

        # Add Row / note
        btn_row = tk.Frame(sec, bg="#ffffff"); btn_row.pack(fill="x", pady=6)
        tk.Button(btn_row, text="➕ Add Row", relief="flat",
                  bg="#2a7d4f", fg="white", font=("Segoe UI", 9, "bold"),
                  padx=10, pady=4, cursor="hand2",
                  command=self._add_row).pack(side="left", padx=4)
        tk.Label(btn_row,
                 text=f"(max {MAX_ROWS} rows)  ·  "
                      "Quote Link: paste a URL or type 'Please see attached quote'",
                 bg="#ffffff", fg="#888", font=("Segoe UI", 8)).pack(
            side="left", padx=8)

    def _add_row(self):
        if len(self._rows) >= MAX_ROWS:
            messagebox.showinfo("Max Rows",
                                f"The template supports a maximum of {MAX_ROWS} line items.")
            return
        idx = len(self._rows)
        bg = "#ffffff" if idx % 2 == 0 else "#f5f7fa"
        row_frame = tk.Frame(self._items_frame, bg=bg, pady=2)
        row_frame.pack(fill="x", pady=1)

        widgets = {}

        # Row number label
        tk.Label(row_frame, text=str(idx + 1), bg=bg,
                 font=("Segoe UI", 9), width=3, anchor="e").pack(side="left", padx=2)

        for col, w in zip(ITEM_COLS, ITEM_WIDTHS):
            col_clean = col.replace("\n", " ")
            if col == "Ext. Price\n(auto)":
                # Read-only calculated field
                var = tk.StringVar(value="")
                e = tk.Entry(row_frame, textvariable=var, width=w,
                             font=("Segoe UI", 9), state="disabled",
                             disabledbackground="#e8eded", disabledforeground="#555",
                             relief="flat")
                e.pack(side="left", padx=2)
                widgets["_ext_var"] = var
                widgets["_ext_entry"] = e
            elif col in ("CP Related\n(Y/N)", "New Supplier\n(Y/N)"):
                var = tk.StringVar(value="N")
                cb = ttk.Combobox(row_frame, textvariable=var,
                                  values=["Y", "N"], width=w - 2,
                                  font=("Segoe UI", 9), state="readonly")
                cb.pack(side="left", padx=2)
                widgets[col_clean] = var
            else:
                var = tk.StringVar()
                e = tk.Entry(row_frame, textvariable=var, width=w,
                             font=("Segoe UI", 9), relief="flat",
                             bg="#fafafa", highlightthickness=1,
                             highlightbackground="#cdd5e0")
                e.pack(side="left", padx=2)
                widgets[col_clean] = var
                # Auto-calc Ext. Price when Qty or Unit Price changes
                if col in ("Qty", "Unit Price"):
                    var.trace_add("write", lambda *_, r=widgets: self._calc_ext(r))

        self._rows.append(widgets)

    def _calc_ext(self, row_widgets):
        try:
            qty = float(row_widgets.get("Qty", tk.StringVar()).get() or 0)
            price = float(row_widgets.get("Unit Price", tk.StringVar()).get() or 0)
            ext = qty * price
            row_widgets["_ext_var"].set(f"{ext:,.2f}" if ext else "")
        except ValueError:
            row_widgets["_ext_var"].set("")

    # ── Footer buttons ─────────────────────────────────────────────────────────
    def _build_footer_buttons(self, parent):
        btn_bar = tk.Frame(parent, bg="#f0f4f8", pady=12)
        btn_bar.pack(fill="x")

        tk.Button(btn_bar, text="🗑  Clear All", relief="flat",
                  bg="#c0392b", fg="white", font=("Segoe UI", 10, "bold"),
                  padx=14, pady=8, cursor="hand2",
                  command=self._clear_all).pack(side="left", padx=6)

        tk.Button(btn_bar, text="💾  Submit & Save to Excel", relief="flat",
                  bg="#1a3c5e", fg="white", font=("Segoe UI", 11, "bold"),
                  padx=18, pady=8, cursor="hand2",
                  command=self._submit).pack(side="right", padx=6)

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _section(self, parent, title):
        outer = tk.LabelFrame(parent, text=title, bg="#ffffff",
                              font=("Segoe UI", 10, "bold"), fg="#1a3c5e",
                              padx=10, pady=8, relief="groove", bd=2)
        outer.pack(fill="x", pady=8)
        return outer

    def _lbl(self, parent, text):
        return tk.Label(parent, text=text, bg="#ffffff",
                        font=("Segoe UI", 9, "bold"), fg="#333")

    def _clear_all(self):
        if not messagebox.askyesno("Clear Form",
                                   "Clear all line items and reset the form?"):
            return
        self.requestor_var.set("")
        self.issuer_var.set(ISSUER)
        self.date_var.set(date.today().strftime("%m/%d/%Y"))
        self.job_id_var.set(JOB_ID_OPTIONS[0])
        for rw in self._rows:
            for key, var in rw.items():
                if key.startswith("_"):
                    continue
                if isinstance(var, tk.StringVar):
                    col = key.replace("\n", " ")
                    if "Y/N" in col:
                        var.set("N")
                    else:
                        var.set("")

    # ── Submit → write Excel ──────────────────────────────────────────────────
    def _submit(self):
        if not OPENPYXL_OK:
            messagebox.showerror("Missing Library",
                                 "openpyxl is not installed.\n"
                                 "Run:  pip install openpyxl")
            return

        # Validate at least one non-empty line item
        has_data = any(
            rw.get("Description", tk.StringVar()).get().strip()
            for rw in self._rows
        )
        if not has_data:
            messagebox.showwarning("No Items",
                                   "Please enter at least one line item description.")
            return

        # Ask where to save
        save_path = filedialog.asksaveasfilename(
            title="Save Purchase Requisition",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=f"PurchaseReq_{date.today().strftime('%Y%m%d')}.xlsx",
        )
        if not save_path:
            return

        try:
            self._write_excel(save_path)
            messagebox.showinfo("Saved",
                                f"Purchase Requisition saved to:\n{save_path}")
        except Exception as exc:
            messagebox.showerror("Save Failed", str(exc))

    def _write_excel(self, save_path):
        # Load template if it exists, otherwise build from scratch
        if os.path.exists(TEMPLATE_PATH):
            wb = openpyxl.load_workbook(TEMPLATE_PATH)
            ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
            self._write_template_skeleton(ws)

        # ── Header cells ────────────────────────────────────────────────────
        ws["B2"] = self.issuer_var.get().strip()
        ws["B3"] = self.date_var.get().strip()
        ws["B4"] = self.requestor_var.get().strip()
        ws["B5"] = SHIP_TO

        job_id = self.job_id_var.get()
        ws["A6"] = f"Job ID Parent :13382-ABA  |  Sub-Job: {job_id}"

        # ── Line items (rows 7-29) ───────────────────────────────────────────
        for row_idx, rw in enumerate(self._rows):
            xl_row = 7 + row_idx
            if xl_row > 29:
                break

            def gv(col):
                return rw.get(col, tk.StringVar()).get().strip()

            ws.cell(xl_row, 2).value  = _safe_num(gv("Qty"))           # B
            ws.cell(xl_row, 3).value  = gv("Description")              # C
            ws.cell(xl_row, 4).value  = gv("Supplier Name")            # D
            ws.cell(xl_row, 5).value  = gv("Part #")                   # E
            ws.cell(xl_row, 7).value  = _safe_num(gv("Unit Price"))    # G
            ws.cell(xl_row, 9).value  = gv("Date Required")            # I
            ws.cell(xl_row, 10).value = gv("CP Related (Y/N)")         # J
            ws.cell(xl_row, 11).value = gv("If YES, CP #")             # K
            ws.cell(xl_row, 12).value = gv("Quote Number")             # L
            ws.cell(xl_row, 13).value = gv("Approved By")              # M
            ws.cell(xl_row, 14).value = gv("Leadtime")                 # N
            ws.cell(xl_row, 15).value = gv("New Supplier (Y/N)")       # O
            ws.cell(xl_row, 16).value = gv("Reason New Supplier Is Needed")  # P

            # Quote Link — F column (col 6)
            quote_raw = gv("Quote Link")
            if quote_raw:
                if _is_url(quote_raw):
                    ws.cell(xl_row, 6).value = quote_raw
                    ws.cell(xl_row, 6).hyperlink = quote_raw
                    ws.cell(xl_row, 6).font = Font(
                        color="0563C1", underline="single")
                else:
                    ws.cell(xl_row, 6).value = quote_raw

            # Ext. Price formula (H)
            qty_addr = f"B{xl_row}"
            uprice_addr = f"G{xl_row}"
            ws.cell(xl_row, 8).value = f"=SUM({qty_addr}*{uprice_addr})"

        # Clear any leftover data from previous saves beyond our rows
        for xl_row in range(7 + len(self._rows), 30):
            for col in range(2, 17):
                if col != 8:  # keep Ext. Price formula
                    ws.cell(xl_row, col).value = None

        wb.save(save_path)

    def _write_template_skeleton(self, ws):
        """Builds header labels in case the template file is missing."""
        ws["A2"] = "Issuer:"
        ws["A3"] = "Date:"
        ws["I3"] = "Purchase Requisition"
        ws["A4"] = "Airbus Requestor:"
        ws["A5"] = "Ship to:"
        ws["A6"] = "Job ID Parent :13382-ABA"
        # Column headers row 6
        for col_num, header in enumerate(
            ["Qty", "Description", "Supplier Name", "Part #",
             "Quote Link", "Unit Price", "Ext. Price",
             "Date Required", "CP Related (Y/N)", "If YES, CP #",
             "Quote Number", "Approved By", "Leadtime",
             "New Supplier (Y/N)", "Reason New Supplier Is Needed"],
            start=2
        ):
            ws.cell(6, col_num).value = header
            ws.cell(6, col_num).font = Font(bold=True)
        # Ext. Price formulas
        for row in range(7, 30):
            ws.cell(row, 8).value = f"=SUM(B{row}*G{row})"
        # Total
        ws["H30"] = "=SUM(H7:H29)"
        ws["A30"] = "Requested By: _________________________________________"
        ws["A31"] = "Authorized By:_________________________________________"


# ── Utility helpers ───────────────────────────────────────────────────────────

def _safe_num(val: str):
    """Return float if val looks numeric, else the raw string."""
    try:
        return float(val)
    except (ValueError, TypeError):
        return val or None


def _is_url(val: str) -> bool:
    return bool(re.match(r"^https?://", val.strip(), re.IGNORECASE))
